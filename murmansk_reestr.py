import openpyxl as op
from datetime import date, timedelta
import json
from data import name_TT_MURM
from openpyxl.formatting.rule import ColorScale, FormatObject
import os 

"""
Datetime 
"""
today = date.today().strftime('%d/%m/%Y')
arrive_day = date.today() + timedelta(days=1)
arrive_day_str = arrive_day.strftime('%d/%m/%Y')

directory = 'D:\code folder'

""" Openpyxl settings """

file_names = {}

workbook_ost = None

for file_name in os.listdir(directory):
    if file_name.endswith('.xlsx'):
        try:
            file_names[file_name] = op.load_workbook(file_name)
        except FileNotFoundError:
            print(f"File {file_name} not found.")
            continue

worksheet_ost = None
worksheet_schema = None
worksheet_miratorg = None
worksheet_Mir_Kolbas = None
worksheet_cherkizovo = None
worksheet_OK = None
for book, value in file_names.items():
    print(book)
    match book:
        case 'Останкино.xlsx':
            workbook_ost = value
            worksheet_ost = workbook_ost['Лист1']
        case 'Схема.xlsx':
            workbook_schema = value
            worksheet_schema = workbook_schema['Лист2']
        case 'Мираторг.xlsx':
            workbook_Miratorg = value
            worksheet_miratorg = workbook_Miratorg['Лист1']
        case 'Мир Колбас.xlsx':
            workbook_Mir_Kolbas = value
            worksheet_Mir_Kolbas = workbook_Mir_Kolbas['Лист1']
        case 'Черкизово.xlsx':
            workbook_cherkizovo = value
            worksheet_cherkizovo = workbook_cherkizovo['Лист1']
        case 'ОК.xlsx':
            workbook_OK = value
            worksheet_OK = workbook_OK['доставка-сборка']
            
#worksheets
column_Mir_Kolbas = worksheet_Mir_Kolbas
column_ost = worksheet_ost
column_mir = worksheet_miratorg
column_OK = worksheet_OK
column_schema = worksheet_schema
column_cherkizovo = worksheet_cherkizovo
row_number = 2

global_schema_index = 2


def Catalog():
    # Адреса Останкино/Мир Колбас
    with open('addresses.json', 'r', encoding='utf-8') as file:
        data = json.load(file)
    row_number = 2
    global_schema_index = 2
    
    # Останкино
    if worksheet_ost != None:
        
        for value in column_ost['G']:
            if value.value:
                value.value = value.value.rstrip()
                # убираем лишние пробелы в конце
            
        for cell_ost in column_ost['G']:
            if cell_ost.value in data:
                day = today
                name = 'Останкино'
                doc_num =  worksheet_ost.cell(row = row_number + 2, column = 3).value
                weight = worksheet_ost.cell(row = row_number + 2, column = 10).value
                temp = worksheet_ost.cell(row = row_number + 2, column = 12).value
                address = cell_ost.value
                for name_TT_OST in name_TT_MURM:
                    if name_TT_OST in address:
                        worksheet_schema.cell(row = global_schema_index, column = 4).value = name_TT_OST  
                worksheet_schema.cell(row = global_schema_index, column = 5, value = address)
                worksheet_schema.cell(row = global_schema_index, column = 7, value = doc_num)
                worksheet_schema.cell(row = global_schema_index, column = 9, value = weight)
                worksheet_schema.cell(row = global_schema_index, column = 6, value = temp)
                worksheet_schema.cell(row = global_schema_index, column = 3, value = name)
                worksheet_schema.cell(row = global_schema_index, column = 1, value = day)
                worksheet_schema.cell(row = global_schema_index, column = 2, value = arrive_day_str)
                row_number += 1
                global_schema_index += 1

            
    # Мираторг
    if worksheet_miratorg != None:
        mira_row_number = 1
        for cell_mir in column_mir['G']:
            if 'Зверосовхоз' in cell_mir.value:
                mira_row_number += 1
                continue
            elif 'Адрес доставки' in cell_mir.value:
                mira_row_number += 1
                continue
            else:
                day = today 
                name = 'Мираторг'
                doc_num = worksheet_miratorg.cell(row = mira_row_number, column = 4).value
                doc_num_no_zero = str(doc_num)[:-4]
                weight = worksheet_miratorg.cell(row = mira_row_number, column = 8).value
                temp = worksheet_miratorg.cell(row = mira_row_number, column = 6).value
                store = worksheet_miratorg.cell(row = mira_row_number, column = 5).value
                if temp == 'Зам':
                    worksheet_schema.cell(row = global_schema_index, column = 6).value = '-18°C'
                else:
                    worksheet_schema.cell(row = global_schema_index, column = 6).value = '0 .. +4°C'
                address = cell_mir.value
                worksheet_schema.cell(row = global_schema_index, column = 4, value = store)
                worksheet_schema.cell(row = global_schema_index, column = 5, value = address)
                worksheet_schema.cell(row = global_schema_index, column = 7, value = doc_num_no_zero)
                worksheet_schema.cell(row = global_schema_index, column = 9, value = weight)
                worksheet_schema.cell(row = global_schema_index, column = 3, value = name)
                worksheet_schema.cell(row = global_schema_index, column = 1, value = day)
                worksheet_schema.cell(row = global_schema_index, column = 2, value = arrive_day_str)
                mira_row_number += 1
                global_schema_index += 1
            
    # Кондитер
    if worksheet_OK != None:
        OK_row_number = 1
        for cell_OK in column_OK['E']:
            if cell_OK.value is None:
                OK_row_number += 1
                continue
            elif cell_OK.value == 'Адрес доставки':
                OK_row_number += 1
                continue
            else:
                day = today
                name = 'Объединенный Кондитер'
                doc_num = worksheet_OK.cell(row = OK_row_number, column = 3).value
                weight = worksheet_OK.cell(row = OK_row_number, column = 6).value
                temp = '0 .. +4°C'
                store = worksheet_OK.cell(row = OK_row_number, column = 4).value
                address = cell_OK.value
                worksheet_schema.cell(row = global_schema_index, column = 5, value = address)
                worksheet_schema.cell(row = global_schema_index, column = 7, value = doc_num)
                worksheet_schema.cell(row = global_schema_index, column = 9, value = weight)
                worksheet_schema.cell(row = global_schema_index, column = 3, value = name)
                worksheet_schema.cell(row = global_schema_index, column = 1, value = day)
                worksheet_schema.cell(row = global_schema_index, column = 2, value = arrive_day_str)
                worksheet_schema.cell(row = global_schema_index, column = 6, value = temp)
                worksheet_schema.cell(row = global_schema_index, column = 4, value = store)
                OK_row_number += 1
                global_schema_index += 1
    
    # Мир Колбас
    if worksheet_Mir_Kolbas != None:
        for value in column_Mir_Kolbas['G']:
            if value.value:
                value.value = value.value.rstrip()
        
        Mir_kolbas_row_number = 2
        for cell_Mir_Kolbas in column_Mir_Kolbas['G']:
            if cell_Mir_Kolbas.value == 'Адрес доставки':
                Mir_kolbas_row_number += 1
                continue 
            elif cell_Mir_Kolbas.value in data:
                day = today 
                name = 'Мир Колбас'
                doc_num = worksheet_Mir_Kolbas.cell(row = Mir_kolbas_row_number + 1, column = 3).value
                weight = worksheet_Mir_Kolbas.cell(row = Mir_kolbas_row_number + 1, column = 10).value 
                temp = '0 .. +4°C'
                address = cell_Mir_Kolbas.value
                for name_TT_MK in name_TT_MURM:
                    if name_TT_MK in address:
                        worksheet_schema.cell(row = global_schema_index, column = 4).value = name_TT_MK
                worksheet_schema.cell(row = global_schema_index, column = 5, value = address)
                worksheet_schema.cell(row = global_schema_index, column = 7, value = doc_num)
                worksheet_schema.cell(row = global_schema_index, column = 9, value = weight)
                worksheet_schema.cell(row = global_schema_index, column = 3, value = name)
                worksheet_schema.cell(row = global_schema_index, column = 1, value = day)
                worksheet_schema.cell(row = global_schema_index, column = 2, value = arrive_day_str)
                worksheet_schema.cell(row = global_schema_index, column = 6, value = temp)
                Mir_kolbas_row_number += 1
                global_schema_index += 1
    
    # Черкизово
    if worksheet_cherkizovo != None:
        Cherkizovo_row_number = 2
        for cell_cherkizovo in column_cherkizovo['H']:
            if cell_cherkizovo.value == 'Адрес назначения':
                Cherkizovo_row_number += 1
                continue
            elif cell_cherkizovo.value in data and cell_cherkizovo.fill.start_color.index == 'FFFFFF00':
                day = today 
                name = 'Черкизово'
                temp = worksheet_cherkizovo.cell(row = Cherkizovo_row_number, column = 13).value
                if temp == '+2 +4':
                    worksheet_schema.cell(row = global_schema_index, column = 6).value = '0 .. +4°C'
                #else:
                    #worksheet_schema.cell(row = global_schema_index, column = 6).value = '-18°C'
                weight = worksheet_cherkizovo.cell(row = Cherkizovo_row_number, column = 11).value
                doc_num = worksheet_cherkizovo.cell(row = Cherkizovo_row_number, column = 3).value
                if doc_num is None:
                    worksheet_schema.cell(row = global_schema_index, column = 7).value = 'Б/Н'
                else:
                    worksheet_schema.cell(row = global_schema_index, column = 7, value = doc_num)
                address = cell_cherkizovo.value
                store = worksheet_cherkizovo.cell(row = Cherkizovo_row_number, column = 7).value
                worksheet_schema.cell(row = global_schema_index, column = 5, value = address)
                worksheet_schema.cell(row = global_schema_index, column = 9, value = weight)
                worksheet_schema.cell(row = global_schema_index, column = 3, value = name)
                worksheet_schema.cell(row = global_schema_index, column = 1, value = day)
                worksheet_schema.cell(row = global_schema_index, column = 2, value = arrive_day_str)
                worksheet_schema.cell(row = global_schema_index, column = 6, value = temp)
                worksheet_schema.cell(row = global_schema_index, column = 4, value = store)
                Cherkizovo_row_number += 1
                global_schema_index += 1
            
    return Catalog



Catalog()


workbook_schema.save('Схема.xlsx')
        